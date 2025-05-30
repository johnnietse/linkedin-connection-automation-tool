import os
import time
import random
import re
import cv2
import numpy as np
import pandas as pd
import openpyxl
import pyautogui
import imutils
from io import BytesIO
from PIL import Image
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (TimeoutException, NoSuchElementException,
                                        WebDriverException, ElementClickInterceptedException)
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
LINKEDIN_EMAIL = os.getenv('LINKEDIN_EMAIL')
LINKEDIN_PASSWORD = os.getenv('LINKEDIN_PASSWORD')


class LinkedInUltimateConnector:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.driver = None
        self.results = []
        self.processed_count = 0
        self.connected_count = 0
        self.skipped_count = 0

        # Paths to template images
        self.connect_templates = [
            "templates/connect_button_1.png",
            "templates/connect_button_2.png",
            "templates/connect_button_3.png"
        ]
        self.send_templates = [
            "templates/send_without_note_1.png",
            "templates/send_without_note_2.png",
            "templates/send_without_note_3.png"
        ]
        self.message_templates = [
            "templates/message_button_1.png",
            "templates/message_button_2.png"
        ]
        self.pending_templates = [
            "templates/pending_button_1.png"
        ]

    def initialize_driver(self):
        """Set up Chrome WebDriver with options to avoid detection"""
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--log-level=3")

        self.driver = webdriver.Chrome(options=options)
        self.driver.set_page_load_timeout(15)
        print("Browser initialized successfully")

    def linkedin_login(self):
        """Log in to LinkedIn account with manual captcha handling"""
        print("Logging in to LinkedIn...")
        self.driver.get("https://www.linkedin.com/login")

        # Wait for manual login
        input(
            "Please complete the login process including any puzzles in the browser. Press ENTER in this console when you're logged in...")

        # Verify successful login
        try:
            WebDriverWait(self.driver, 15).until(
                lambda d: "feed" in d.current_url or d.find_elements(By.ID, "global-nav")
            )
            print("Login successful")
            return True
        except Exception as e:
            print(f"Login verification failed: {str(e)}")
            return False

    def extract_hyperlinks(self):
        """Extract URLs from Excel with improved handling"""
        print(f"Extracting hyperlinks from: {self.excel_path}")
        wb = openpyxl.load_workbook(self.excel_path)
        sheet = wb.active

        urls = []
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=1)
            url = None

            # 1. Check for hyperlink
            if cell.hyperlink:
                url = cell.hyperlink.target

            # 2. Check cell value
            elif cell.value:
                value = str(cell.value).strip()
                # Clean common URL prefixes
                if value.startswith("www."):
                    value = "https://" + value
                # Extract URL from text
                if "linkedin.com" in value:
                    url = value
                elif "http" in value:
                    url = value

            # 3. Skip header if needed
            if row == 1 and url and ("linkedin.com" not in url.lower()):
                print(f"Skipping header row: {url}")
                continue

            if url:
                # Ensure proper formatting
                if not url.startswith("http"):
                    url = "https://" + url
                urls.append(url)
            else:
                print(f"Skipping row {row}: '{cell.value}'")

        print(f"Extracted {len(urls)} URLs")
        return urls

    def get_window_info(self):
        """Get current window position and size"""
        try:
            window_x = self.driver.execute_script("return window.screenX;")
            window_y = self.driver.execute_script("return window.screenY;")
            window_width = self.driver.execute_script("return window.outerWidth;")
            window_height = self.driver.execute_script("return window.outerHeight;")
            return window_x, window_y, window_width, window_height
        except:
            # Fallback to full screen
            return 0, 0, pyautogui.size().width, pyautogui.size().height

    def capture_screenshot(self):
        """Capture screenshot of browser window"""
        try:
            window_x, window_y, window_width, window_height = self.get_window_info()
            screenshot = pyautogui.screenshot(region=(window_x, window_y, window_width, window_height))
            return np.array(screenshot)
        except Exception as e:
            print(f"Screenshot capture error: {str(e)}")
            return None

    def multi_scale_template_match(self, screenshot, template_path, min_confidence=0.65):
        """Match template at multiple scales"""
        if screenshot is None:
            return None, 0, 1.0

        try:
            screenshot_cv = cv2.cvtColor(screenshot, cv2.COLOR_RGB2BGR)
            gray_screenshot = cv2.cvtColor(screenshot_cv, cv2.COLOR_BGR2GRAY)

            template = cv2.imread(template_path)
            if template is None:
                return None, 0, 1.0

            gray_template = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)

            best_match_val = 0
            best_match_loc = None
            best_scale = 1.0

            # Try multiple scales
            for scale in np.linspace(0.7, 1.3, 5):
                resized_template = imutils.resize(gray_template, width=int(gray_template.shape[1] * scale))

                if resized_template.shape[0] > gray_screenshot.shape[0] or resized_template.shape[1] > \
                        gray_screenshot.shape[1]:
                    continue

                result = cv2.matchTemplate(gray_screenshot, resized_template, cv2.TM_CCOEFF_NORMED)
                _, max_val, _, max_loc = cv2.minMaxLoc(result)

                if max_val > best_match_val and max_val > min_confidence:
                    best_match_val = max_val
                    best_match_loc = max_loc
                    best_scale = scale

            return best_match_loc, best_match_val, best_scale
        except Exception as e:
            print(f"Template matching error: {str(e)}")
            return None, 0, 1.0

    def find_button_by_text(self, text, element_type="button", partial_match=False):
        """Find button by visible text using XPath"""
        try:
            if partial_match:
                xpath = f"//{element_type}[contains(., '{text}')]"
            else:
                xpath = f"//{element_type}[text()='{text}']"

            elements = self.driver.find_elements(By.XPATH, xpath)

            # Filter visible elements
            visible_elements = [el for el in elements if el.is_displayed()]

            if visible_elements:
                return visible_elements[0]
            return None
        except Exception:
            return None

    def find_connect_button(self):
        """Prioritize finding Connect button above all else"""
        # First try text matching (exact match)
        connect_button = self.find_button_by_text("Connect")
        if connect_button:
            return connect_button, "text_exact"

        # Try text matching (partial match)
        connect_button = self.find_button_by_text("Connect", partial_match=True)
        if connect_button:
            return connect_button, "text_partial"

        # Try image recognition as fallback
        window_x, window_y, window_width, window_height = self.get_window_info()
        screenshot = self.capture_screenshot()

        best_button = None
        best_confidence = 0

        for template_path in self.connect_templates:
            if not os.path.exists(template_path):
                continue

            # Focus on top-right quadrant
            roi = screenshot[0:window_height // 2, window_width // 2:window_width]
            match_loc, match_val, match_scale = self.multi_scale_template_match(roi, template_path)

            if match_val > 0.65 and match_val > best_confidence:
                template = cv2.imread(template_path)
                h, w = template.shape[:2]
                scaled_w = int(w * match_scale)
                scaled_h = int(h * match_scale)

                center_x = window_x + window_width // 2 + match_loc[0] + scaled_w // 2
                center_y = window_y + match_loc[1] + scaled_h // 2

                best_button = (center_x, center_y)
                best_confidence = match_val

        if best_button:
            return best_button, "image"

        return None, "not_found"

    def find_send_button(self):
        """Find Send without a note button with priority"""
        # First try text matching (exact match)
        send_button = self.find_button_by_text("Send without a note")
        if send_button:
            return send_button, "text_exact"

        # Try text matching (partial match)
        send_button = self.find_button_by_text("Send without a note", partial_match=True)
        if not send_button:
            send_button = self.find_button_by_text("Send without", partial_match=True)
        if send_button:
            return send_button, "text_partial"

        # Try image recognition as fallback
        window_x, window_y, window_width, window_height = self.get_window_info()
        screenshot = self.capture_screenshot()

        best_button = None
        best_confidence = 0

        for template_path in self.send_templates:
            if not os.path.exists(template_path):
                continue

            # Focus on center of screen
            roi = screenshot[window_height // 4:3 * window_height // 4, window_width // 4:3 * window_width // 4]
            match_loc, match_val, match_scale = self.multi_scale_template_match(roi, template_path, min_confidence=0.60)

            if match_val > 0.60 and match_val > best_confidence:
                template = cv2.imread(template_path)
                h, w = template.shape[:2]
                scaled_w = int(w * match_scale)
                scaled_h = int(h * match_scale)

                center_x = window_x + window_width // 4 + match_loc[0] + scaled_w // 2
                center_y = window_y + window_height // 4 + match_loc[1] + scaled_h // 2

                best_button = (center_x, center_y)
                best_confidence = match_val

        if best_button:
            return best_button, "image"

        return None, "not_found"

    def is_connected_state(self):
        """Check for Message or Pending state using hybrid approach"""
        # First try text matching
        if self.find_button_by_text("Message") or self.find_button_by_text("Pending"):
            return True

        # Then try image recognition
        screenshot = self.capture_screenshot()
        window_x, window_y, window_width, window_height = self.get_window_info()

        # Check for Message buttons
        for template_path in self.message_templates:
            if os.path.exists(template_path):
                # Focus on top-right quadrant
                roi = screenshot[0:window_height // 2, window_width // 2:window_width]
                match_loc, match_val, _ = self.multi_scale_template_match(roi, template_path, min_confidence=0.60)
                if match_val > 0.60:
                    return True

        # Check for Pending buttons
        for template_path in self.pending_templates:
            if os.path.exists(template_path):
                # Focus on top-right quadrant
                roi = screenshot[0:window_height // 2, window_width // 2:window_width]
                match_loc, match_val, _ = self.multi_scale_template_match(roi, template_path, min_confidence=0.60)
                if match_val > 0.60:
                    return True

        return False

    def click_element(self, element, element_type):
        """Click element based on its type"""
        try:
            if element_type in ["text_exact", "text_partial"]:
                # DOM element - scroll into view and click with JavaScript
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                self.driver.execute_script("arguments[0].click();", element)
                return True
            elif element_type == "image":
                # Coordinates - click with pyautogui
                pyautogui.moveTo(element[0], element[1], duration=random.uniform(0.1, 0.3))
                pyautogui.click()
                return True
        except Exception as e:
            print(f"Click error: {str(e)}")
            return False

    def send_connection_request(self):
        """Click Connect and Send buttons with absolute priority"""
        try:
            # 1. Check if already connected (skip if true)
            if self.is_connected_state():
                print(" - Already connected or pending")
                return "Already Connected"

            # 2. Find and click Connect button
            connect_element, connect_type = self.find_connect_button()
            if not connect_element:
                print(" - Connect button not found")
                return "Connect Button Not Found"

            print(f" - Connect button found via {connect_type}")
            if not self.click_element(connect_element, connect_type):
                print(" - Failed to click Connect button")
                return "Click Failed"

            # Short delay for modal to appear
            time.sleep(1 + random.uniform(0.3, 0.7))

            # 3. Find and click Send button
            send_element, send_type = self.find_send_button()
            if not send_element:
                print(" - Send button not found, assuming direct connection")
                return True  # Connection still sent

            print(f" - Send button found via {send_type}")
            if not self.click_element(send_element, send_type):
                print(" - Failed to click Send button")
                return "Click Failed"

            return True

        except Exception as e:
            print(f"Connection error: {str(e)}")
            return f"Error: {str(e)}"

    def process_profile(self, url):
        """Process a single LinkedIn profile"""
        start_time = time.time()
        try:
            # Navigate to profile with reduced timeout
            try:
                self.driver.get(url)
            except TimeoutException:
                print(" - Page load timed out, continuing anyway")

            # Short initial wait
            time.sleep(0.7 + random.uniform(0.2, 0.5))

            # Scroll to make buttons visible
            self.driver.execute_script("window.scrollTo(0, 400);")

            # Attempt to send connection
            result = self.send_connection_request()

            if result is True:
                self.connected_count += 1
                return "Connection Sent"
            elif result == "Already Connected":
                self.skipped_count += 1
                return result
            else:
                return result

        except WebDriverException as e:
            if "ERR_NAME_NOT_RESOLVED" in str(e):
                return "Invalid URL"
            return f"Error: {str(e)}"
        except Exception as e:
            return f"Error: {str(e)}"
        finally:
            elapsed = time.time() - start_time
            print(f" - Processed in {elapsed:.1f} seconds")

    def save_progress(self):
        """Save results to Excel"""
        try:
            df = pd.DataFrame(self.results)
            df.to_excel('data/results.xlsx', index=False)
            print(f"Progress saved: {len(self.results)} profiles processed")
        except Exception as e:
            print(f"Error saving progress: {str(e)}")

    def run(self):
        """Main execution flow"""
        try:
            # Extract URLs
            urls = self.extract_hyperlinks()
            total = len(urls)

            if total == 0:
                print("No valid URLs found")
                return

            print(f"Found {total} LinkedIn URLs to process")

            # Initialize browser and login
            self.initialize_driver()
            if not self.linkedin_login():
                return

            # Process all URLs
            for i, url in enumerate(urls):
                self.processed_count += 1
                print(f"\nProcessing ({self.processed_count}/{total}): {url}")

                # Process profile
                status = self.process_profile(url)
                self.results.append({
                    'URL': url,
                    'Status': status,
                    'Timestamp': pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
                })

                # Save progress periodically
                if self.processed_count % 10 == 0:
                    self.save_progress()

                # Calculate delay
                base_delay = random.uniform(6, 12)
                delay = max(4, base_delay)
                print(f"Waiting {delay:.1f} seconds...")
                time.sleep(delay)

                # Take breaks
                if self.connected_count > 0 and self.connected_count % 20 == 0:
                    nap = random.randint(120, 240)  # 2-4 minutes
                    print(f"⏸️ Taking {nap // 60} min break...")
                    time.sleep(nap)

            print("\nProcessing complete!")

        except KeyboardInterrupt:
            print("\nProcess interrupted by user")
        except Exception as e:
            print(f"Critical error: {str(e)}")
        finally:
            # Final save and cleanup
            self.save_progress()
            if self.driver:
                self.driver.quit()
                print("Browser closed")

            # Print summary
            print(f"\n{'=' * 50}")
            print(f"Total profiles processed: {self.processed_count}")
            print(f"Connection requests sent: {self.connected_count}")
            print(f"Skipped profiles: {self.skipped_count}")
            print(f"Results saved to: data/results.xlsx")
            print(f"{'=' * 50}")


if __name__ == "__main__":
    # Configuration
    INPUT_FILE = "data/input_profiles.xlsx"

    # Create required directories
    os.makedirs("data", exist_ok=True)
    os.makedirs("templates", exist_ok=True)

    # Verify template images exist
    required_templates = [
        "connect_button_1.png", "connect_button_2.png", "connect_button_3.png",
        "send_without_note_1.png", "send_without_note_2.png", "send_without_note_3.png",
        "message_button_1.png", "message_button_2.png",
        "pending_button_1.png"
    ]
    missing = [t for t in required_templates if not os.path.exists(f"templates/{t}")]

    if missing:
        print(f"Warning: Missing template images: {', '.join(missing)}")
        print("Text-based detection will be used primarily")

    # Run the automation
    automator = LinkedInUltimateConnector(INPUT_FILE)
    automator.run()